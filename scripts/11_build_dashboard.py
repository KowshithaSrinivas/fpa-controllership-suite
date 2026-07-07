from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
BOLD = Font(name=FONT, bold=True)
TITLE_FONT = Font(name=FONT, size=16, bold=True, color="1F4E78")
KPI_FONT = Font(name=FONT, size=20, bold=True, color="1F4E78")
KPI_LABEL_FONT = Font(name=FONT, size=10, color="595959")
CARD_FILL = PatternFill("solid", start_color="F2F2F2")
GREEN = Font(name=FONT, color="008000")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%'
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
MONTH_LABELS = [pd.Period(m).strftime("%b-%y") for m in MONTHS]

wb = load_workbook(OUT)
ws = wb.create_sheet("Executive_Dashboard", 0)  # insert as first visible tab after Cover conceptually
wb.move_sheet("Executive_Dashboard", offset=-(len(wb.sheetnames) - 2))
ws.sheet_view.showGridLines = False

ws["B2"] = "Executive Dashboard — Consolidated (All 3 Entities)"
ws["B2"].font = TITLE_FONT
ws["B3"] = "FY2025 | Source: Microsoft Dynamics 365 F&O (simulated) | Auto-refreshes from P&L / Balance Sheet / Cost Center tabs"
ws["B3"].font = Font(name=FONT, italic=True, size=9, color="808080")

# ---- KPI cards row ----
kpis = [
    ("FY Revenue", f"='P&L'!O8"),
    ("FY EBITDA", f"='P&L'!O26"),
    ("EBITDA Margin", f"='P&L'!O26/'P&L'!O8"),
    ("FY Net Income", f"='P&L'!O32"),
    ("Dec-25 Cash (Selected Entity, see P&L tab)", "=Cash_Flow!M23"),
]
# NOTE: 'P&L' col O = consolidated FY total column (col 15). Confirm mapping below.
col_start = 2
for i, (label, formula) in enumerate(kpis):
    c = col_start + i * 3
    for cc in range(c, c + 2):
        for rr in range(5, 8):
            ws.cell(row=rr, column=cc).fill = CARD_FILL
            ws.cell(row=rr, column=cc).border = BORDER
    ws.cell(row=5, column=c, value=label).font = KPI_LABEL_FONT
    val_cell = ws.cell(row=6, column=c, value=formula)
    val_cell.font = KPI_FONT
    val_cell.number_format = PCT_FMT if "Margin" in label else CURRENCY_FMT
    ws.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + 1)
    ws.merge_cells(start_row=6, start_column=c, end_row=7, end_column=c + 1)

# ---- Helper data block for charts (feeds off P&L / Cost Center via formulas) ----
HELPER_ROW = 10
ws.cell(row=HELPER_ROW, column=2, value="Chart Data (hidden helper rows below; formulas reference other tabs)").font = Font(name=FONT, italic=True, size=9, color="808080")

# Revenue trend (consolidated) - pull month by month using SUMIFS directly on Data_GL (no entity filter = consolidated)
r_rev = HELPER_ROW + 1
ws.cell(row=r_rev, column=1, value="Month")
ws.cell(row=r_rev + 1, column=1, value="Revenue")
ws.cell(row=r_rev + 2, column=1, value="EBITDA Margin %")
for i, per in enumerate(MONTHS):
    col = 2 + i
    ws.cell(row=r_rev, column=col, value=MONTH_LABELS[i])
    rev_f = f'=SUMIFS(Data_GL!$G:$G,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"4000")+SUMIFS(Data_GL!$G:$G,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"4010")'
    ws.cell(row=r_rev + 1, column=col, value=rev_f).number_format = CURRENCY_FMT
    cogs_f = f'SUMIFS(Data_GL!$G:$G,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"5000")+SUMIFS(Data_GL!$G:$G,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"5010")'
    opex_f = "+".join([f'SUMIFS(Data_GL!$G:$G,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"{a}")' for a in ["6000","6010","6020","6030","6040"]])
    colL = get_column_letter(col)
    ws.cell(row=r_rev + 2, column=col,
            value=f"=IF({colL}{r_rev+1}=0,0,({colL}{r_rev+1}-({cogs_f})-({opex_f}))/{colL}{r_rev+1})")
    ws.cell(row=r_rev + 2, column=col).number_format = PCT_FMT

# Budget vs actual bridge (consolidated, full year, for revenue + opex)
r_bva = r_rev + 4
ws.cell(row=r_bva, column=1, value="Bridge Item")
ws.cell(row=r_bva + 1, column=1, value="FY Actual Revenue")
ws.cell(row=r_bva + 2, column=1, value="FY Budget Revenue")
ws.cell(row=r_bva + 1, column=2, value='=SUMIFS(Data_GL!$G:$G,Data_GL!$C:$C,"4000")+SUMIFS(Data_GL!$G:$G,Data_GL!$C:$C,"4010")').number_format = CURRENCY_FMT
ws.cell(row=r_bva + 2, column=2, value='=SUMIFS(Data_Budget!$F:$F,Data_Budget!$C:$C,"4000")+SUMIFS(Data_Budget!$F:$F,Data_Budget!$C:$C,"4010")').number_format = CURRENCY_FMT

# Cost center spend split (consolidated, FY, OpEx only)
r_cc = r_bva + 4
cc_list = [("CC100", "Sales & Marketing"), ("CC200", "Operations"), ("CC300", "G&A"),
           ("CC400", "R&D / Product"), ("UNALLOCATED", "Unallocated")]
ws.cell(row=r_cc, column=1, value="Cost Center")
ws.cell(row=r_cc, column=2, value="FY OpEx Spend")
for i, (code, name) in enumerate(cc_list):
    ws.cell(row=r_cc + 1 + i, column=1, value=name)
    f = f'=SUMIFS(Data_GL!$G:$G,Data_GL!$F:$F,"{code}",Data_GL!$E:$E,"OpEx")'
    ws.cell(row=r_cc + 1 + i, column=2, value=f).number_format = CURRENCY_FMT

wb.save(OUT)
print("Dashboard data blocks built. r_rev", r_rev, "r_bva", r_bva, "r_cc", r_cc)
