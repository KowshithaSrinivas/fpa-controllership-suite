import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

CLEAN = "/home/claude/fpa-project/data/clean"
OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", start_color="DCE6F1")
YELLOW = PatternFill("solid", start_color="FFFF00")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
MONTH_LABELS = [pd.Period(m).strftime("%b-%y") for m in MONTHS]
ENTITIES = ["DE01", "DE02", "US01"]
ENTITY_NAMES = {"DE01": "Germany GmbH Berlin", "DE02": "Germany GmbH Munich", "US01": "US Holdings Inc"}

tb = pd.read_csv(f"{CLEAN}/gl_trial_balance_clean.csv")
bs = pd.read_csv(f"{CLEAN}/balance_sheet_clean.csv")
bud = pd.read_csv(f"{CLEAN}/budget_clean.csv")
coa = pd.read_csv(f"{CLEAN}/chart_of_accounts.csv")
cc = pd.read_csv(f"{CLEAN}/cost_centers.csv")
dq = pd.read_csv(f"{CLEAN}/data_quality_log.csv")

wb = Workbook()
wb.remove(wb.active)

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# =========================================================
# COVER SHEET
# =========================================================
ws = wb.create_sheet("Cover")
ws["B2"] = "Multi-Entity FP&A & Controllership Reporting Suite"
ws["B2"].font = Font(name=FONT, size=18, bold=True, color="1F4E78")
ws["B4"] = "Simulated month-end close, budgeting, forecasting, and management reporting package"
ws["B4"].font = Font(name=FONT, size=12, italic=True)
ws["B6"] = "Source system: Microsoft Dynamics 365 Finance & Operations (simulated exports)"
ws["B6"].font = Font(name=FONT, size=11, bold=True)

info = [
    ("Entities covered", "DE01 - Germany GmbH Berlin | DE02 - Germany GmbH Munich | US01 - US Holdings Inc"),
    ("Reporting currency", "USD (EUR entities translated at month-end FX rate)"),
    ("Period", "Jan 2025 - Dec 2025"),
    ("", ""),
    ("Tab", "Purpose"),
    ("Data_GL / Data_BS / Data_Budget", "Cleaned, validated data pulled from D365 exports (source of all formulas below)"),
    ("P&L", "Monthly P&L by entity + consolidated, fully formula-driven"),
    ("Balance_Sheet", "Monthly balance sheet by entity + consolidated, with balance check"),
    ("Cash_Flow", "Indirect-method cash flow statement derived from P&L and Balance Sheet"),
    ("Budget_vs_Actual", "Monthly variance analysis with commentary flags"),
    ("Rolling_Forecast", "6-month rolling forecast off trailing trend, adjustable growth assumption"),
    ("Cost_Center_Drilldown", "Spend by cost center with over/under-budget flags for business partnering"),
    ("Month_End_Close_Checklist", "Close task tracker (owner, due date, status) incl. data-quality log from D365 import"),
    ("Controls_Checklist", "Internal controls / segregation-of-duties tracker"),
    ("Executive_Dashboard", "KPI summary and charts for management presentation"),
]
r = 8
for label, val in info:
    ws.cell(row=r, column=2, value=label).font = BOLD if val else Font(name=FONT, bold=True, size=12)
    ws.cell(row=r, column=3, value=val).font = BLACK
    r += 1
set_widths(ws, {"A": 3, "B": 32, "C": 90})
ws.sheet_view.showGridLines = False

# =========================================================
# DATA_GL (hidden-ish helper source data — kept visible for transparency)
# =========================================================
ws = wb.create_sheet("Data_GL")
cols = ["LegalEntity", "Period", "MainAccount", "AccountName", "AccountCategory", "CostCenter", "AmountUSD"]
ws.append(cols)
style_header_row(ws, 1, len(cols))
for _, row in tb.iterrows():
    ws.append([row[c] for c in cols])
set_widths(ws, {"A": 12, "B": 10, "C": 12, "D": 26, "E": 14, "F": 14, "G": 14})
GL_LAST_ROW = ws.max_row

ws_bs = wb.create_sheet("Data_BS")
cols_bs = ["LegalEntity", "Period", "MainAccount", "AccountName", "AccountCategory", "AmountUSD"]
ws_bs.append(cols_bs)
style_header_row(ws_bs, 1, len(cols_bs))
for _, row in bs.iterrows():
    ws_bs.append([row[c] for c in cols_bs])
set_widths(ws_bs, {"A": 12, "B": 10, "C": 12, "D": 26, "E": 14, "F": 14})
BS_LAST_ROW = ws_bs.max_row

ws_bud = wb.create_sheet("Data_Budget")
cols_bud = ["LegalEntity", "Period", "MainAccount", "AccountName", "AccountCategory", "BudgetAmount"]
ws_bud.append(cols_bud)
style_header_row(ws_bud, 1, len(cols_bud))
for _, row in bud.iterrows():
    ws_bud.append([row[c] for c in cols_bud])
set_widths(ws_bud, {"A": 12, "B": 10, "C": 12, "D": 26, "E": 14, "F": 14})
BUD_LAST_ROW = ws_bud.max_row

wb.save(OUT)
print("Phase 1 saved: Cover + Data sheets. GL rows:", GL_LAST_ROW, "BS rows:", BS_LAST_ROW, "Budget rows:", BUD_LAST_ROW)
