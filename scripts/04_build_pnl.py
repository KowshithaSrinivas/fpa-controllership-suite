from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", start_color="DCE6F1")
YELLOW = PatternFill("solid", start_color="FFFF00")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
MONTH_LABELS = [pd.Period(m).strftime("%b-%y") for m in MONTHS]
N_MONTHS = len(MONTHS)

wb = load_workbook(OUT)
ws = wb.create_sheet("P&L")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

ws["B2"] = "Select Entity:"
ws["B2"].font = BOLD
ws["C2"] = "DE01"
ws["C2"].font = BLUE
ws["C2"].fill = YELLOW
dv = DataValidation(type="list", formula1='"DE01,DE02,US01"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws["C2"])
ws["D2"] = "<- change entity to view; Consolidated column always sums all 3 entities"
ws["D2"].font = Font(name=FONT, italic=True, size=9, color="808080")

HEADER_ROW = 4
COL_FIRST_MONTH = 2
COL_FY = COL_FIRST_MONTH + N_MONTHS
COL_CONS = COL_FY + 1
LAST_COL = COL_CONS

ws.cell(row=HEADER_ROW, column=1, value="P&L Line Item")
for i, lab in enumerate(MONTH_LABELS):
    ws.cell(row=HEADER_ROW, column=COL_FIRST_MONTH + i, value=lab)
ws.cell(row=HEADER_ROW, column=COL_FY, value="FY Total (Selected Entity)")
ws.cell(row=HEADER_ROW, column=COL_CONS, value="FY Total (Consolidated)")
for c in range(1, LAST_COL + 1):
    cell = ws.cell(row=HEADER_ROW, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

r = HEADER_ROW + 1
ROW = {}
def add_row(key, label, bold=False, shade=False):
    global r
    ws.cell(row=r, column=1, value=label)
    if bold:
        ws.cell(row=r, column=1).font = BOLD
    if shade:
        for c in range(1, LAST_COL + 1):
            ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
    ROW[key] = r
    r += 1

add_row("hdr_rev", "Revenue", bold=True)
add_row("4000", "  Product Sales (4000)")
add_row("4010", "  Services (4010)")
add_row("total_rev", "Total Revenue", bold=True, shade=True)
r += 1
add_row("hdr_cogs", "Cost of Goods Sold", bold=True)
add_row("5000", "  Materials (5000)")
add_row("5010", "  Freight (5010)")
add_row("total_cogs", "Total COGS", bold=True, shade=True)
r += 1
add_row("gp", "Gross Profit", bold=True, shade=True)
add_row("gp_pct", "Gross Margin %")
r += 1
add_row("hdr_opex", "Operating Expenses", bold=True)
add_row("6000", "  Salaries & Wages (6000)")
add_row("6010", "  Marketing (6010)")
add_row("6020", "  IT & Software (6020)")
add_row("6030", "  Facilities & Rent (6030)")
add_row("6040", "  Professional Fees (6040)")
add_row("total_opex", "Total OpEx (before D&A)", bold=True, shade=True)
r += 1
add_row("ebitda", "EBITDA", bold=True, shade=True)
add_row("ebitda_pct", "EBITDA Margin %")
r += 1
add_row("7000", "Depreciation & Amortization (7000)")
add_row("ebit", "EBIT", bold=True, shade=True)
add_row("8000", "Interest Expense (8000)")
add_row("ni", "Net Income", bold=True, shade=True)
add_row("ni_pct", "Net Margin %")

ACCOUNT_ROWS = ["4000", "4010", "5000", "5010", "6000", "6010", "6020", "6030", "6040", "7000", "8000"]

for acc in ACCOUNT_ROWS:
    rr = ROW[acc]
    for i, per in enumerate(MONTHS):
        col = COL_FIRST_MONTH + i
        f_ent = f'=SUMIFS(Data_GL!$G:$G,Data_GL!$A:$A,$C$2,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"{acc}")'
        c = ws.cell(row=rr, column=col, value=f_ent)
        c.font = GREEN
        c.number_format = CURRENCY_FMT
    colL_s, colL_e = get_column_letter(COL_FIRST_MONTH), get_column_letter(COL_FIRST_MONTH + N_MONTHS - 1)
    fy = ws.cell(row=rr, column=COL_FY, value=f"=SUM({colL_s}{rr}:{colL_e}{rr})")
    fy.number_format = CURRENCY_FMT
    cons = ws.cell(row=rr, column=COL_CONS, value=f'=SUMIFS(Data_GL!$G:$G,Data_GL!$C:$C,"{acc}")')
    cons.font = GREEN
    cons.number_format = CURRENCY_FMT

def write_formula_row(key, build_formula):
    rr = ROW[key]
    for col in range(COL_FIRST_MONTH, LAST_COL + 1):
        colL = get_column_letter(col)
        cell = ws.cell(row=rr, column=col, value=build_formula(colL))
        cell.number_format = CURRENCY_FMT

write_formula_row("total_rev", lambda c: f"={c}{ROW['4000']}+{c}{ROW['4010']}")
write_formula_row("total_cogs", lambda c: f"={c}{ROW['5000']}+{c}{ROW['5010']}")
write_formula_row("gp", lambda c: f"={c}{ROW['total_rev']}-{c}{ROW['total_cogs']}")
write_formula_row("total_opex", lambda c: (
    f"={c}{ROW['6000']}+{c}{ROW['6010']}+{c}{ROW['6020']}+{c}{ROW['6030']}+{c}{ROW['6040']}"))
write_formula_row("ebitda", lambda c: f"={c}{ROW['gp']}-{c}{ROW['total_opex']}")
write_formula_row("ebit", lambda c: f"={c}{ROW['ebitda']}-{c}{ROW['7000']}")
write_formula_row("ni", lambda c: f"={c}{ROW['ebit']}-{c}{ROW['8000']}")

for key, numer_key in [("gp_pct", "gp"), ("ebitda_pct", "ebitda"), ("ni_pct", "ni")]:
    rr = ROW[key]
    for col in range(COL_FIRST_MONTH, LAST_COL + 1):
        colL = get_column_letter(col)
        cell = ws.cell(row=rr, column=col,
                        value=f"=IF({colL}{ROW['total_rev']}=0,0,{colL}{ROW[numer_key]}/{colL}{ROW['total_rev']})")
        cell.number_format = PCT_FMT
        cell.font = Font(name=FONT, italic=True)

widths = {"A": 34}
for i in range(N_MONTHS):
    widths[get_column_letter(COL_FIRST_MONTH + i)] = 11
widths[get_column_letter(COL_FY)] = 16
widths[get_column_letter(COL_CONS)] = 16
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT)
import json
print("P&L sheet built. ROW map:", json.dumps(ROW))
