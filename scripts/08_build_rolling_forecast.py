from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
FORECAST_FILL = PatternFill("solid", start_color="FFF2CC")
YELLOW = PatternFill("solid", start_color="FFFF00")
SUBTOTAL_FILL = PatternFill("solid", start_color="DCE6F1")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'

ACTUAL_MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
ACTUAL_LABELS = [pd.Period(m).strftime("%b-%y") for m in ACTUAL_MONTHS]
FORECAST_MONTHS = pd.date_range("2026-01-01", "2026-06-01", freq="MS").strftime("%Y-%m").tolist()
FORECAST_LABELS = [pd.Period(m).strftime("%b-%y") for m in FORECAST_MONTHS]
ALL_LABELS = ACTUAL_LABELS + FORECAST_LABELS
N_ACT = len(ACTUAL_MONTHS)
N_FCST = len(FORECAST_MONTHS)

wb = load_workbook(OUT)
ws = wb.create_sheet("Rolling_Forecast")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B6"

ws["B2"] = "Select Entity:"; ws["B2"].font = BOLD
ws["C2"] = "DE01"; ws["C2"].font = BLUE; ws["C2"].fill = YELLOW
dv = DataValidation(type="list", formula1='"DE01,DE02,US01"', allow_blank=False)
ws.add_data_validation(dv); dv.add(ws["C2"])

ws["B3"] = "Monthly Growth Assumption (applied to forecast months):"; ws["B3"].font = BOLD
ws["E3"] = 0.015; ws["E3"].font = BLUE; ws["E3"].fill = YELLOW
ws["E3"].number_format = "0.0%"
ws["F3"] = "<- editable assumption (default trailing-3mo avg growth rate)"
ws["F3"].font = Font(name=FONT, italic=True, size=9, color="808080")

HEADER_ROW = 5
COL_FIRST = 2
ws.cell(row=HEADER_ROW, column=1, value="Revenue Forecast Line")
for i, lab in enumerate(ALL_LABELS):
    c = ws.cell(row=HEADER_ROW, column=COL_FIRST + i, value=lab)
for c in range(1, COL_FIRST + N_ACT + N_FCST):
    cell = ws.cell(row=HEADER_ROW, column=c)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

ROW = {"label": HEADER_ROW}
r = HEADER_ROW + 1
ws.cell(row=r, column=1, value="Total Revenue (Actual / Forecast)").font = BOLD
REV_ROW = r
r += 1
ws.cell(row=r, column=1, value="Trailing 3-Month Avg Growth Rate").font = Font(name=FONT, italic=True)
GROWTH_ROW = r
r += 1
ws.cell(row=r, column=1, value="Variance vs prior Rolling Forecast (n/a for actuals)").font = Font(name=FONT, italic=True)
VAR_ROW = r

# Actual revenue months: pull from Data_GL
for i, per in enumerate(ACTUAL_MONTHS):
    col = COL_FIRST + i
    f = f'=SUMIFS(Data_GL!$G:$G,Data_GL!$A:$A,$C$2,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"4000")+SUMIFS(Data_GL!$G:$G,Data_GL!$A:$A,$C$2,Data_GL!$B:$B,"{per}",Data_GL!$C:$C,"4010")'
    cell = ws.cell(row=REV_ROW, column=col, value=f)
    cell.font = GREEN
    cell.number_format = CURRENCY_FMT

# Forecast months: prior month * (1+growth assumption)
for i in range(N_FCST):
    col = COL_FIRST + N_ACT + i
    prevL = get_column_letter(col - 1)
    cell = ws.cell(row=REV_ROW, column=col, value=f"={prevL}{REV_ROW}*(1+$E$3)")
    cell.number_format = CURRENCY_FMT
    cell.fill = FORECAST_FILL
    ws.cell(row=HEADER_ROW, column=col).fill = PatternFill("solid", start_color="BF8F00")

# Trailing 3-month growth (only meaningful from month 4 onward on actuals, shown across all for reference)
for i in range(3, N_ACT + N_FCST):
    col = COL_FIRST + i
    colL = get_column_letter(col)
    l3 = get_column_letter(col - 3)
    cell = ws.cell(row=GROWTH_ROW, column=col, value=f"=IF({l3}{REV_ROW}=0,0,({colL}{REV_ROW}-{l3}{REV_ROW})/{l3}{REV_ROW}/3)")
    cell.number_format = "0.0%"

ws.column_dimensions["A"].width = 40
for i in range(N_ACT + N_FCST):
    ws.column_dimensions[get_column_letter(COL_FIRST + i)].width = 11

wb.save(OUT)
print("Rolling Forecast built. REV_ROW", REV_ROW, "GROWTH_ROW", GROWTH_ROW)
