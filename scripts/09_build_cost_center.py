from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
YELLOW = PatternFill("solid", start_color="FFFF00")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()

COST_CENTERS = [
    ("CC100", "Sales & Marketing"),
    ("CC200", "Operations"),
    ("CC300", "G&A"),
    ("CC400", "R&D / Product"),
    ("UNALLOCATED", "Unallocated / Needs Review"),
]

wb = load_workbook(OUT)
ws = wb.create_sheet("Cost_Center_Drilldown")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A6"

ws["B2"] = "Select Entity:"; ws["B2"].font = BOLD
ws["C2"] = "DE01"; ws["C2"].font = BLUE; ws["C2"].fill = YELLOW
dv1 = DataValidation(type="list", formula1='"DE01,DE02,US01"', allow_blank=False)
ws.add_data_validation(dv1); dv1.add(ws["C2"])

ws["B3"] = "Select Period:"; ws["B3"].font = BOLD
ws["C3"] = "2025-06"; ws["C3"].font = BLUE; ws["C3"].fill = YELLOW
dv2 = DataValidation(type="list", formula1='"' + ",".join(MONTHS) + '"', allow_blank=False)
ws.add_data_validation(dv2); dv2.add(ws["C3"])

HEADER_ROW = 5
headers = ["Cost Center", "Actual Spend", "Budget (Total OpEx, allocated pro-rata)", "% of Total Spend", "Variance %", "Business Partnering Flag"]
for i, h in enumerate(headers):
    cell = ws.cell(row=HEADER_ROW, column=1 + i, value=h)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

r = HEADER_ROW + 1
first_r = r
for cc_code, cc_name in COST_CENTERS:
    ws.cell(row=r, column=1, value=f"{cc_name} ({cc_code})")
    f = f'=SUMIFS(Data_GL!$G:$G,Data_GL!$A:$A,$C$2,Data_GL!$B:$B,$C$3,Data_GL!$F:$F,"{cc_code}",Data_GL!$E:$E,"OpEx")'
    ac = ws.cell(row=r, column=2, value=f); ac.font = GREEN; ac.number_format = CURRENCY_FMT
    r += 1
last_r = r - 1
total_row = r
ws.cell(row=total_row, column=1, value="Total OpEx Spend").font = BOLD
ws.cell(row=total_row, column=2, value=f"=SUM(B{first_r}:B{last_r})").number_format = CURRENCY_FMT
ws.cell(row=total_row, column=2).font = BOLD

# Budget allocated pro-rata to cost center based on its share of total OpEx budget for the entity/period
budget_total_f = f'SUMIFS(Data_Budget!$F:$F,Data_Budget!$A:$A,$C$2,Data_Budget!$B:$B,$C$3,Data_Budget!$E:$E,"OpEx")'
for i, (cc_code, cc_name) in enumerate(COST_CENTERS):
    rr = first_r + i
    share_f = f"=IF($B${total_row}=0,0,B{rr}/$B${total_row})*({budget_total_f})"
    bc = ws.cell(row=rr, column=3, value=share_f)
    bc.number_format = CURRENCY_FMT
    ws.cell(row=rr, column=4, value=f"=IF($B${total_row}=0,0,B{rr}/$B${total_row})").number_format = PCT_FMT
    ws.cell(row=rr, column=5, value=f"=IF(C{rr}=0,0,(B{rr}-C{rr})/C{rr})").number_format = PCT_FMT
    ws.cell(row=rr, column=6, value=f'=IF(E{rr}>0.15,"Over budget - discuss with cost center owner",IF(E{rr}<-0.15,"Under-spending - confirm on track",""))')

rng = f"E{first_r}:E{last_r}"
ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0.15"], fill=RED_FILL))
ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-0.15"], fill=GREEN_FILL))

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 26
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 36

wb.save(OUT)
print("Cost Center Drilldown built.")
