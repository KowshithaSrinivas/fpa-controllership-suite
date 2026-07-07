from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", start_color="DCE6F1")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
MONTH_LABELS = [pd.Period(m).strftime("%b-%y") for m in MONTHS]
N_MONTHS = len(MONTHS)

wb = load_workbook(OUT)
ws = wb.create_sheet("Cash_Flow")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

ws["B2"] = "Indirect method | Entity follows the selector on the P&L tab (cell P&L!$C$2)"
ws["B2"].font = Font(name=FONT, italic=True, size=10, color="808080")

HEADER_ROW = 4
COL_FIRST = 2
LAST_COL = COL_FIRST + N_MONTHS - 1

ws.cell(row=HEADER_ROW, column=1, value="Cash Flow Line Item")
for i, lab in enumerate(MONTH_LABELS):
    ws.cell(row=HEADER_ROW, column=COL_FIRST + i, value=lab)
for c in range(1, LAST_COL + 1):
    cell = ws.cell(row=HEADER_ROW, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

r = HEADER_ROW + 1
ROW = {}
def add_row(key, label, bold=False, shade=False):
    global r
    ws.cell(row=r, column=1, value=label)
    if bold:
        ws.cell(row=r, column=1).font = BOLD
    if shade:
        for c in range(1, LAST_COL + 1):
            ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
    ROW[key] = r
    r += 1

add_row("hdr_op", "Operating Activities", bold=True)
add_row("ni", "  Net Income (from P&L)")
add_row("da", "  + Depreciation & Amortization (non-cash)")
add_row("d_ar", "  (Increase)/Decrease in Accounts Receivable")
add_row("d_ap", "  Increase/(Decrease) in Accounts Payable")
add_row("d_accr", "  Increase/(Decrease) in Accrued Liabilities")
add_row("cfo", "Cash Flow from Operations", bold=True, shade=True)
r += 1
add_row("hdr_inv", "Investing Activities", bold=True)
add_row("capex", "  Capital Expenditures (proxy: PP&E build)")
add_row("cfi", "Cash Flow from Investing", bold=True, shade=True)
r += 1
add_row("hdr_fin", "Financing Activities", bold=True)
add_row("debt_chg", "  Debt Issuance/(Repayment)")
add_row("cff", "Cash Flow from Financing", bold=True, shade=True)
r += 1
add_row("net_chg", "Net Change in Cash", bold=True, shade=True)
add_row("beg_cash", "  Beginning Cash Balance")
add_row("end_cash", "  Ending Cash Balance (check vs Balance_Sheet)", bold=True, shade=True)

PNL_ROW = {"ni": 32, "7000": 29}  # from P&L sheet build (row map)
BS_ROW = {"1200": 7, "2000": 12, "2100": 13, "1500": 8, "2500": 14, "1000": 6}  # from Balance_Sheet row map

for i, per in enumerate(MONTHS):
    col = COL_FIRST + i
    colL = get_column_letter(col)
    prevL = get_column_letter(col - 1) if i > 0 else None

    ws.cell(row=ROW["ni"], column=col, value=f"='P&L'!{colL}{PNL_ROW['ni']}").font = GREEN
    ws.cell(row=ROW["da"], column=col, value=f"='P&L'!{colL}{PNL_ROW['7000']}").font = GREEN

    if i == 0:
        ws.cell(row=ROW["d_ar"], column=col, value=0)
        ws.cell(row=ROW["d_ap"], column=col, value=0)
        ws.cell(row=ROW["d_accr"], column=col, value=0)
        ws.cell(row=ROW["capex"], column=col, value=0)
        ws.cell(row=ROW["debt_chg"], column=col, value=0)
        ws.cell(row=ROW["beg_cash"], column=col, value=f"=Balance_Sheet!{colL}{BS_ROW['1000']}-{colL}{ROW['net_chg']}")
    else:
        ws.cell(row=ROW["d_ar"], column=col,
                 value=f"=-(Balance_Sheet!{colL}{BS_ROW['1200']}-Balance_Sheet!{prevL}{BS_ROW['1200']})").font = GREEN
        ws.cell(row=ROW["d_ap"], column=col,
                 value=f"=Balance_Sheet!{colL}{BS_ROW['2000']}-Balance_Sheet!{prevL}{BS_ROW['2000']}").font = GREEN
        ws.cell(row=ROW["d_accr"], column=col,
                 value=f"=Balance_Sheet!{colL}{BS_ROW['2100']}-Balance_Sheet!{prevL}{BS_ROW['2100']}").font = GREEN
        ws.cell(row=ROW["capex"], column=col,
                 value=f"=-(Balance_Sheet!{colL}{BS_ROW['1500']}-Balance_Sheet!{prevL}{BS_ROW['1500']})-{colL}{ROW['da']}").font = GREEN
        ws.cell(row=ROW["debt_chg"], column=col,
                 value=f"=Balance_Sheet!{colL}{BS_ROW['2500']}-Balance_Sheet!{prevL}{BS_ROW['2500']}").font = GREEN
        ws.cell(row=ROW["beg_cash"], column=col, value=f"={prevL}{ROW['end_cash']}")

    ws.cell(row=ROW["cfo"], column=col,
            value=f"={colL}{ROW['ni']}+{colL}{ROW['da']}+{colL}{ROW['d_ar']}+{colL}{ROW['d_ap']}+{colL}{ROW['d_accr']}")
    ws.cell(row=ROW["cfi"], column=col, value=f"={colL}{ROW['capex']}")
    ws.cell(row=ROW["cff"], column=col, value=f"={colL}{ROW['debt_chg']}")
    ws.cell(row=ROW["net_chg"], column=col,
            value=f"={colL}{ROW['cfo']}+{colL}{ROW['cfi']}+{colL}{ROW['cff']}")
    ws.cell(row=ROW["end_cash"], column=col, value=f"={colL}{ROW['beg_cash']}+{colL}{ROW['net_chg']}")

for rr in ROW.values():
    for col in range(COL_FIRST, LAST_COL + 1):
        ws.cell(row=rr, column=col).number_format = CURRENCY_FMT

widths = {"A": 42}
for i in range(N_MONTHS):
    widths[get_column_letter(COL_FIRST + i)] = 11
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT)
import json
print("Cash Flow built. ROW map:", json.dumps(ROW))
