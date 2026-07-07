from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
YELLOW = PatternFill("solid", start_color="FFFF00")
RED_FILL = PatternFill("solid", start_color="FFC7CE")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
MONTH_LABELS = [pd.Period(m).strftime("%b-%y") for m in MONTHS]

ACCOUNTS = [
    ("4000", "Product Sales", "Revenue"),
    ("4010", "Services", "Revenue"),
    ("5000", "Materials (COGS)", "COGS"),
    ("5010", "Freight (COGS)", "COGS"),
    ("6000", "Salaries & Wages", "OpEx"),
    ("6010", "Marketing", "OpEx"),
    ("6020", "IT & Software", "OpEx"),
    ("6030", "Facilities & Rent", "OpEx"),
    ("6040", "Professional Fees", "OpEx"),
]

wb = load_workbook(OUT)
ws = wb.create_sheet("Budget_vs_Actual")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A6"

ws["B2"] = "Select Entity:"; ws["B2"].font = BOLD
ws["C2"] = "DE01"; ws["C2"].font = BLUE; ws["C2"].fill = YELLOW
dv1 = DataValidation(type="list", formula1='"DE01,DE02,US01"', allow_blank=False)
ws.add_data_validation(dv1); dv1.add(ws["C2"])

ws["B3"] = "Select Period:"; ws["B3"].font = BOLD
ws["C3"] = "2025-06"; ws["C3"].font = BLUE; ws["C3"].fill = YELLOW
dv2 = DataValidation(type="list", formula1='"' + ",".join(MONTHS) + '"', allow_blank=False)
ws.add_data_validation(dv2); dv2.add(ws["C3"])

HEADER_ROW = 5
headers = ["Account", "Category", "Actual", "Budget", "Variance $", "Variance %", "Commentary Flag"]
for i, h in enumerate(headers):
    cell = ws.cell(row=HEADER_ROW, column=1 + i, value=h)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

r = HEADER_ROW + 1
for acc, name, cat in ACCOUNTS:
    ws.cell(row=r, column=1, value=f"{name} ({acc})")
    ws.cell(row=r, column=2, value=cat)
    actual_f = f'=SUMIFS(Data_GL!$G:$G,Data_GL!$A:$A,$C$2,Data_GL!$B:$B,$C$3,Data_GL!$C:$C,"{acc}")'
    budget_f = f'=SUMIFS(Data_Budget!$F:$F,Data_Budget!$A:$A,$C$2,Data_Budget!$B:$B,$C$3,Data_Budget!$C:$C,"{acc}")'
    ac = ws.cell(row=r, column=3, value=actual_f); ac.font = GREEN; ac.number_format = CURRENCY_FMT
    bc = ws.cell(row=r, column=4, value=budget_f); bc.font = GREEN; bc.number_format = CURRENCY_FMT
    # For revenue: variance = Actual - Budget (favorable if positive). For cost lines: variance = Budget - Actual (favorable if positive, i.e. under budget)
    if cat == "Revenue":
        var_f = f"=C{r}-D{r}"
    else:
        var_f = f"=D{r}-C{r}"
    vc = ws.cell(row=r, column=5, value=var_f); vc.number_format = CURRENCY_FMT
    pc = ws.cell(row=r, column=6, value=f"=IF(D{r}=0,0,E{r}/ABS(D{r}))"); pc.number_format = PCT_FMT
    ws.cell(row=r, column=7, value=f'=IF(F{r}<-0.1,"Investigate - Unfavorable >10%",IF(F{r}>0.1,"Favorable - review driver",""))')
    r += 1

TOTAL_ROW = r
ws.cell(row=TOTAL_ROW, column=1, value="Total (this view)").font = BOLD
for col in [3, 4, 5]:
    colL = get_column_letter(col)
    ws.cell(row=TOTAL_ROW, column=col, value=f"=SUM({colL}{HEADER_ROW+1}:{colL}{TOTAL_ROW-1})").number_format = CURRENCY_FMT
    ws.cell(row=TOTAL_ROW, column=col).font = BOLD

rng = f"F{HEADER_ROW+1}:F{TOTAL_ROW-1}"
ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-0.1"], fill=RED_FILL))
ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0.1"], fill=GREEN_FILL))

ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 12
for c in ["C", "D", "E"]:
    ws.column_dimensions[c].width = 14
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 30

wb.save(OUT)
print("Budget vs Actual built.")
