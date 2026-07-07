from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
GREEN = Font(name=FONT, color="008000")
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill("solid", start_color="DCE6F1")
YELLOW = PatternFill("solid", start_color="FFFF00")
OK_FILL = PatternFill("solid", start_color="C6EFCE")
CURRENCY_FMT = '$#,##0;($#,##0);"-"'

MONTHS = pd.date_range("2025-01-01", "2025-12-01", freq="MS").strftime("%Y-%m").tolist()
MONTH_LABELS = [pd.Period(m).strftime("%b-%y") for m in MONTHS]
N_MONTHS = len(MONTHS)

wb = load_workbook(OUT)
ws = wb.create_sheet("Balance_Sheet")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B5"

ws["B2"] = "Select Entity:"
ws["B2"].font = BOLD
ws["C2"] = "DE01"
ws["C2"].font = BLUE
ws["C2"].fill = YELLOW
dv = DataValidation(type="list", formula1='"DE01,DE02,US01"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws["C2"])

HEADER_ROW = 4
COL_FIRST = 2
COL_CONS = COL_FIRST + N_MONTHS
LAST_COL = COL_CONS

ws.cell(row=HEADER_ROW, column=1, value="Balance Sheet Line Item")
for i, lab in enumerate(MONTH_LABELS):
    ws.cell(row=HEADER_ROW, column=COL_FIRST + i, value=lab)
ws.cell(row=HEADER_ROW, column=COL_CONS, value="Dec-25 (Consolidated)")
for c in range(1, LAST_COL + 1):
    cell = ws.cell(row=HEADER_ROW, column=c)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

r = HEADER_ROW + 1
ROW = {}
def add_row(key, label, bold=False, shade=False):
    global r
    ws.cell(row=r, column=1, value=label)
    if bold:
        ws.cell(row=r, column=1).font = BOLD
    if shade:
        for c in range(1, LAST_COL + 1):
            ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
    ROW[key] = r
    r += 1

add_row("hdr_assets", "Assets", bold=True)
add_row("1000", "  Cash & Cash Equivalents")
add_row("1200", "  Accounts Receivable")
add_row("1500", "  Fixed Assets, net")
add_row("total_assets", "Total Assets", bold=True, shade=True)
r += 1
add_row("hdr_liab", "Liabilities", bold=True)
add_row("2000", "  Accounts Payable")
add_row("2100", "  Accrued Liabilities")
add_row("2500", "  Long-term Debt")
add_row("total_liab", "Total Liabilities", bold=True, shade=True)
r += 1
add_row("hdr_eq", "Equity", bold=True)
add_row("3000", "  Share Capital")
add_row("3900", "  Retained Earnings")
add_row("total_eq", "Total Equity", bold=True, shade=True)
r += 1
add_row("total_le", "Total Liabilities + Equity", bold=True, shade=True)
add_row("check", "Balance Check (Assets - L&E, should be 0)", bold=True)

ACCOUNTS = ["1000", "1200", "1500", "2000", "2100", "2500", "3000", "3900"]

# For BS, only Dec-25 closing balances feed the consolidated column (point-in-time, not summed across months)
for acc in ACCOUNTS:
    rr = ROW[acc]
    for i, per in enumerate(MONTHS):
        col = COL_FIRST + i
        f = f'=SUMIFS(Data_BS!$F:$F,Data_BS!$A:$A,$C$2,Data_BS!$B:$B,"{per}",Data_BS!$C:$C,"{acc}")'
        c = ws.cell(row=rr, column=col, value=f)
        c.font = GREEN
        c.number_format = CURRENCY_FMT
    cons = ws.cell(row=rr, column=COL_CONS,
                    value=f'=SUMIFS(Data_BS!$F:$F,Data_BS!$B:$B,"2025-12",Data_BS!$C:$C,"{acc}")')
    cons.font = GREEN
    cons.number_format = CURRENCY_FMT

def write_row(key, build_formula):
    rr = ROW[key]
    for col in range(COL_FIRST, LAST_COL + 1):
        colL = get_column_letter(col)
        cell = ws.cell(row=rr, column=col, value=build_formula(colL))
        cell.number_format = CURRENCY_FMT

write_row("total_assets", lambda c: f"={c}{ROW['1000']}+{c}{ROW['1200']}+{c}{ROW['1500']}")
write_row("total_liab", lambda c: f"={c}{ROW['2000']}+{c}{ROW['2100']}+{c}{ROW['2500']}")
write_row("total_eq", lambda c: f"={c}{ROW['3000']}+{c}{ROW['3900']}")
write_row("total_le", lambda c: f"={c}{ROW['total_liab']}+{c}{ROW['total_eq']}")
write_row("check", lambda c: f"=ROUND({c}{ROW['total_assets']}-{c}{ROW['total_le']},0)")

# Conditional formatting: balance check row green if 0
from openpyxl.formatting.rule import CellIsRule
check_row = ROW["check"]
rng = f"{get_column_letter(COL_FIRST)}{check_row}:{get_column_letter(LAST_COL)}{check_row}"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"], fill=OK_FILL))

widths = {"A": 34}
for i in range(N_MONTHS):
    widths[get_column_letter(COL_FIRST + i)] = 11
widths[get_column_letter(COL_CONS)] = 18
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT)
import json
print("Balance Sheet built. ROW map:", json.dumps(ROW))
