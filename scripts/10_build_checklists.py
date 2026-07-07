from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import pandas as pd

OUT = "/home/claude/fpa-project/model/FPA_Controllership_Suite.xlsx"
CLEAN = "/home/claude/fpa-project/data/clean"
FONT = "Arial"
BOLD = Font(name=FONT, bold=True)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")
RED_FILL = PatternFill("solid", start_color="FFC7CE")

wb = load_workbook(OUT)

# ============ Month-End Close Checklist ============
ws = wb.create_sheet("Month_End_Close_Checklist")
ws.sheet_view.showGridLines = False
ws["B2"] = "Month-End Close Checklist"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color="1F4E78")
ws["B3"] = "Coordinated with Shared Services; owners below are illustrative roles, not names."
ws["B3"].font = Font(name=FONT, italic=True, size=9, color="808080")

headers = ["Task", "Owner", "Due (Business Day)", "Status", "Notes"]
HEADER_ROW = 5
for i, h in enumerate(headers):
    c = ws.cell(row=HEADER_ROW, column=2 + i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

tasks = [
    ("Lock prior period in D365 GL", "Controller", "BD1", "Complete", "Prevents backdated postings"),
    ("Pull GL trial balance export from D365 F&O", "FP&A Analyst", "BD1", "Complete", "Source: General ledger > Trial balance"),
    ("Validate D365 export (duplicates, blank dimensions, FX)", "FP&A Analyst", "BD2", "Complete", "See Data_Quality_Log tab below"),
    ("Confirm accruals with Shared Services (AP/AR)", "Shared Services Lead", "BD2", "In Progress", "Awaiting confirmation from 1 of 3 entities"),
    ("Reconcile intercompany balances", "Controller", "BD3", "In Progress", ""),
    ("Post month-end journal entries", "Controller", "BD3", "Not Started", ""),
    ("Update P&L / Balance Sheet / Cash Flow model", "FP&A Analyst", "BD4", "Not Started", "This workbook"),
    ("Run Budget vs Actual variance analysis", "FP&A Analyst", "BD4", "Not Started", ""),
    ("Business partnering review with cost center owners", "FP&A Analyst", "BD5", "Not Started", "Flag >15% variances"),
    ("Prepare Executive Dashboard / management pack", "FP&A Analyst", "BD5", "Not Started", ""),
    ("Management review & sign-off", "Finance Director", "BD6", "Not Started", ""),
    ("Distribute final reporting package", "Controller", "BD6", "Not Started", ""),
]
r = HEADER_ROW + 1
first_r = r
for t in tasks:
    for i, val in enumerate(t):
        ws.cell(row=r, column=2 + i, value=val)
    r += 1
last_r = r - 1
dv = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,Blocked"', allow_blank=False)
ws.add_data_validation(dv)
for rr in range(first_r, last_r + 1):
    dv.add(ws.cell(row=rr, column=5))
status_rng = f"E{first_r}:E{last_r}"
ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Complete"'], fill=GREEN_FILL))
ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"In Progress"'], fill=YELLOW_FILL))
ws.conditional_formatting.add(status_rng, CellIsRule(operator="equal", formula=['"Blocked"'], fill=RED_FILL))

r += 2
ws.cell(row=r, column=2, value="Data Quality Log (from D365 export validation, feeds this close cycle)").font = BOLD
r += 1
dq_header_row = r
dq = pd.read_csv(f"{CLEAN}/data_quality_log.csv")
for i, col in enumerate(dq.columns):
    c = ws.cell(row=dq_header_row, column=2 + i, value=col)
    c.font = HEADER_FONT; c.fill = HEADER_FILL
r += 1
for _, row in dq.iterrows():
    for i, col in enumerate(dq.columns):
        ws.cell(row=r, column=2 + i, value=row[col])
    r += 1

ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 42

# ============ Controls Checklist ============
ws2 = wb.create_sheet("Controls_Checklist")
ws2.sheet_view.showGridLines = False
ws2["B2"] = "Internal Controls / Compliance Checklist"
ws2["B2"].font = Font(name=FONT, size=14, bold=True, color="1F4E78")
ws2["B3"] = "Controller acts as compliance ambassador / gatekeeper of internal controls"
ws2["B3"].font = Font(name=FONT, italic=True, size=9, color="808080")

headers2 = ["Control", "Frequency", "Owner", "Segregation of Duties Check", "Status"]
for i, h in enumerate(headers2):
    c = ws2.cell(row=5, column=2 + i, value=h)
    c.font = HEADER_FONT; c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", wrap_text=True)

controls = [
    ("Journal entries >$10k require second approver", "Every posting", "Controller", "Preparer ≠ Approver", "Active"),
    ("Bank reconciliation reviewed and signed off", "Monthly", "Controller", "Preparer ≠ Reviewer", "Active"),
    ("D365 user access review (finance module)", "Quarterly", "Controller / IT", "Access matches role", "Active"),
    ("Vendor master changes require dual approval", "Every change", "AP Lead", "Requestor ≠ Approver", "Active"),
    ("Intercompany balances reconciled and eliminated", "Monthly", "Controller", "n/a", "Active"),
    ("3rd-party tax/accounting provider deliverables reviewed", "Monthly", "Controller", "Provider ≠ Reviewer", "Active"),
    ("Budget vs Actual variances >10% documented", "Monthly", "FP&A Analyst", "n/a", "Active"),
]
r2 = 6
for c_ in controls:
    for i, val in enumerate(c_):
        ws2.cell(row=r2, column=2 + i, value=val)
    r2 += 1

ws2.column_dimensions["B"].width = 46
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 26
ws2.column_dimensions["F"].width = 12

wb.save(OUT)
print("Close checklist + controls checklist built.")
